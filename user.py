import openpyxl
import datetime
from openpyxl.styles import Font
from datetime import datetime
import os


now = datetime.now() #DateTime now
dt_string = now.strftime("%m/%d/%Y")# The format DateTime now

#Function to convert string to datetime
def convert(date_time): 
    format = '%m/%d/%Y' # The format
    datetime_str = datetime.strptime(date_time, format) 
    return datetime_str

#Function get Date expire from command "net user rpaadmin002 /domain"(Substring only date) and count remain date
def substring_date(Lines):
    count = 0
    for line in Lines:
        count += 1
        if count == 12:
          T = line.strip()
          date = T[29:]
    date_st = date.split(" ")[0:][0] #subsring only date
    date_expire = convert(date_st) #convert from string to datetime
    coutdate = date_expire-date_now_convert
    coutdate1 = str(coutdate) #Convert from datetime to string
    coutdate1 = coutdate1.split(" ")[0:][0] #substring only count date remain
    return coutdate1

#Command get detail user
wb = openpyxl.load_workbook('user.xlsx')
sheet = wb['Serverinfo']

date_now_convert = convert(dt_string) #convert from string to datetime now
count_row = 13

for rowNum in range(14, sheet.max_row):
    userID = sheet.cell(row=rowNum, column=13).value
    
    if userID==None:
        wb.save('user.xlsx')
        break
    if userID!="-":
        os.system("net user " + str(userID) + " /domain > user.txt")
        f = open("user.txt","r")
        Line = f.readlines()
        remaindate=substring_date(Line)
        remainTest = sheet.cell(row=rowNum, column=12).value

        update_remain = sheet.cell(row=rowNum, column=13).value
        print(update_remain)
        print(remaindate)
        print("--------------")
        if update_remain==userID:
           sheet.cell(row=rowNum, column=12).value = remaindate
           if int(remaindate)<=int("8"):
              sheet.cell(row=rowNum, column=12).font = Font(color="00FF0000")
           else:
              sheet.cell(row=rowNum, column=12).font = Font(color="00000000")
    if userID=="-":
        sheet.cell(row=rowNum, column=12).value = "-"
